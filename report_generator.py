"""
report_generator.py
Генерирует стилизованный Excel-отчёт по статистике Telegram-канала.
4 листа: Сводка / Подписчики / Посты / По часам
"""

import os
import tempfile
from datetime import datetime, timedelta
from typing import Optional

import aiosqlite
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# ПАЛИТРА
# ══════════════════════════════════════════════════════════════════════════════

C_ACCENT   = "5B6EF5"
C_ACCENT2  = "7C3AED"
C_POSITIVE = "22C55E"
C_NEGATIVE = "EF4444"
C_NEUTRAL  = "F59E0B"
C_DARK     = "0F1117"
C_HEADER   = "1E2235"
C_WHITE    = "FFFFFF"
C_MID      = "94A3B8"
C_ROW_ODD  = "F8FAFC"
C_ROW_EVEN = "FFFFFF"
C_BORDER   = "E2E8F0"
FONT       = "Arial"


def _fill(c):    return PatternFill("solid", fgColor=c)
def _font(size=10, bold=False, color=C_DARK, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
def _border(c=C_BORDER):
    s = Side(border_style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bottom(c):
    return Border(bottom=Side(border_style="medium", color=c))


def _header_row(ws, row, cols, col_start=1):
    for i, text in enumerate(cols, col_start):
        c = ws.cell(row, i, text)
        c.font, c.fill, c.border, c.alignment = (
            _font(9, True, C_WHITE), _fill(C_HEADER), _border(C_HEADER), _align()
        )
    ws.row_dimensions[row].height = 20


def _data_row(ws, row, vals, col_start=1, odd=True):
    bg = C_ROW_ODD if odd else C_ROW_EVEN
    for i, v in enumerate(vals, col_start):
        c = ws.cell(row, i, v)
        c.fill, c.border, c.alignment, c.font = (
            _fill(bg), _border(), _align(), _font(9)
        )
    ws.row_dimensions[row].height = 18


def _widths(ws, d: dict):
    for col, w in d.items():
        ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════════
# ДАННЫЕ ИЗ БД
# ══════════════════════════════════════════════════════════════════════════════

async def _fetch(db_path: str, channel_id: int, period_days: Optional[int]) -> dict:
    since = (
        (datetime.utcnow() - timedelta(days=period_days)).isoformat()
        if period_days else "2000-01-01"
    )

    async with aiosqlite.connect(db_path) as db:
        async with db.execute(
            """SELECT DATE(taken_at), MIN(members), MAX(members), MAX(members)-MIN(members)
               FROM snapshots WHERE channel_id=? AND taken_at>=?
               GROUP BY DATE(taken_at) ORDER BY DATE(taken_at)""",
            (channel_id, since),
        ) as cur:
            sub_daily = await cur.fetchall()

        async with db.execute(
            "SELECT members FROM snapshots WHERE channel_id=? ORDER BY taken_at DESC LIMIT 1",
            (channel_id,),
        ) as cur:
            row = await cur.fetchone()
            current = row[0] if row else 0

        async with db.execute(
            "SELECT members FROM snapshots WHERE channel_id=? AND taken_at>=? ORDER BY taken_at LIMIT 1",
            (channel_id, since),
        ) as cur:
            row = await cur.fetchone()
            start = row[0] if row else current

        async with db.execute(
            """SELECT DATE(posted_at), COUNT(*), SUM(has_media),
                      SUM(views), MAX(views), AVG(views)
               FROM posts WHERE channel_id=? AND posted_at>=?
               GROUP BY DATE(posted_at) ORDER BY DATE(posted_at)""",
            (channel_id, since),
        ) as cur:
            posts_daily = await cur.fetchall()

        async with db.execute(
            """SELECT CAST(strftime('%H', posted_at) AS INTEGER), COUNT(*)
               FROM posts WHERE channel_id=? AND posted_at>=?
               GROUP BY 1 ORDER BY 1""",
            (channel_id, since),
        ) as cur:
            hourly = await cur.fetchall()

        async with db.execute(
            "SELECT COUNT(*), SUM(has_media), SUM(views), MAX(views), AVG(views) FROM posts WHERE channel_id=? AND posted_at>=?",
            (channel_id, since),
        ) as cur:
            row = await cur.fetchone()
            posts_total   = row[0] or 0
            media_total   = row[1] or 0
            views_sum     = int(row[2] or 0)
            views_max     = int(row[3] or 0)
            views_avg     = round(row[4] or 0)

        async with db.execute(
            """SELECT DATE(taken_at), MAX(members)-MIN(members)
               FROM snapshots WHERE channel_id=? AND taken_at>=?
               GROUP BY DATE(taken_at) ORDER BY 2 DESC LIMIT 1""",
            (channel_id, since),
        ) as cur:
            best_sub_day = await cur.fetchone()

        async with db.execute(
            """SELECT DATE(taken_at), MAX(members)-MIN(members)
               FROM snapshots WHERE channel_id=? AND taken_at>=?
               GROUP BY DATE(taken_at) ORDER BY 2 ASC LIMIT 1""",
            (channel_id, since),
        ) as cur:
            worst_sub_day = await cur.fetchone()

        async with db.execute(
            """SELECT CAST(strftime('%H', posted_at) AS INTEGER), COUNT(*)
               FROM posts WHERE channel_id=? AND posted_at>=?
               GROUP BY 1 ORDER BY 2 DESC LIMIT 1""",
            (channel_id, since),
        ) as cur:
            best_hour = await cur.fetchone()

    growth     = current - start
    growth_pct = round(growth / start * 100, 2) if start else 0

    return dict(
        current=current, start=start, growth=growth, growth_pct=growth_pct,
        posts_total=posts_total, media_total=media_total,
        views_sum=views_sum, views_max=views_max, views_avg=views_avg,
        best_sub_day=best_sub_day, worst_sub_day=worst_sub_day, best_hour=best_hour,
        sub_daily=sub_daily, posts_daily=posts_daily, hourly=hourly,
    )


# ══════════════════════════════════════════════════════════════════════════════
# ЛИСТ 1 — СВОДКА
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_summary(wb: Workbook, channel: str, period: str, d: dict, ts: str):
    ws = wb.active
    ws.title = "Сводка"

    # Баннер
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = f"📊  АНАЛИТИКА КАНАЛА  ·  {channel.upper()}"
    c.font      = _font(17, True, C_WHITE)
    c.fill      = _fill(C_ACCENT)
    c.alignment = _align()
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = f"Период: {period}     ·     Сформирован: {ts} UTC"
    c.font      = _font(9, italic=True, color=C_MID)
    c.fill      = _fill(C_DARK)
    c.alignment = _align()
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 10

    sign  = "+" if d["growth"] >= 0 else ""
    gcol  = C_POSITIVE if d["growth"] >= 0 else C_NEGATIVE

    kpis = [
        ("👥 ПОДПИСЧИКОВ",   f"{d['current']:,}",                              C_ACCENT),
        ("📈 ПРИРОСТ",        f"{sign}{d['growth']:,} ({sign}{d['growth_pct']}%)", gcol),
        ("✍️ ПОСТОВ",         f"{d['posts_total']:,}",                          C_ACCENT2),
        ("👁 ПРОСМОТРОВ",    f"{d['views_sum']:,}",                             C_NEUTRAL),
        ("🚀 ЛУЧШИЙ ПОСТ",   f"{d['views_max']:,} просм.",                      C_POSITIVE),
        ("📊 СРЕДНЕЕ/ПОСТ",  f"{d['views_avg']:,} просм.",                      C_MID),
    ]

    positions = [(4,"A"),(4,"C"),(4,"E"),(8,"A"),(8,"C"),(8,"E")]

    for (row, col_l), (title, value, color) in zip(positions, kpis):
        col = ord(col_l) - 64
        ecl = get_column_letter(col + 1)

        ws.merge_cells(f"{col_l}{row}:{ecl}{row}")
        t = ws[f"{col_l}{row}"]
        t.value, t.font, t.fill, t.alignment = title, _font(8, True, C_WHITE), _fill(color), _align()
        ws.row_dimensions[row].height = 20

        ws.merge_cells(f"{col_l}{row+1}:{ecl}{row+1}")
        v = ws[f"{col_l}{row+1}"]
        v.value, v.font, v.fill, v.alignment = value, _font(13, True, color), _fill(C_DARK), _align()
        ws.row_dimensions[row+1].height = 30

        for cc in [col, col+1]:
            ws.cell(row+2, cc).border = _bottom(color)
            ws.cell(row+2, cc).fill   = _fill(C_DARK)
        ws.row_dimensions[row+2].height = 4

    ws.row_dimensions[12].height = 14

    # Инсайты
    tr = 13
    ws.merge_cells(f"A{tr}:H{tr}")
    c = ws[f"A{tr}"]
    c.value, c.font, c.fill, c.alignment = (
        "КЛЮЧЕВЫЕ ИНСАЙТЫ", _font(10, True, C_WHITE), _fill(C_HEADER), _align()
    )
    ws.row_dimensions[tr].height = 22

    insights = []
    if d["best_sub_day"] and d["best_sub_day"][1] and d["best_sub_day"][1] > 0:
        insights.append(("🌟 Лучший день роста", f"{d['best_sub_day'][0]}  (+{d['best_sub_day'][1]} подп.)", C_POSITIVE))
    if d["worst_sub_day"] and d["worst_sub_day"][1] and d["worst_sub_day"][1] < 0:
        insights.append(("📉 Наибольшая убыль", f"{d['worst_sub_day'][0]}  ({d['worst_sub_day'][1]} подп.)", C_NEGATIVE))
    if d["best_hour"]:
        h = d["best_hour"][0]
        insights.append(("⏰ Лучшее время постов", f"{h:02d}:00 – {h:02d}:59 UTC", C_NEUTRAL))
    if d["posts_total"]:
        mp = round(d["media_total"] / d["posts_total"] * 100)
        insights.append(("🖼 Постов с медиа", f"{d['media_total']} из {d['posts_total']}  ({mp}%)", C_ACCENT))
    if not insights:
        insights.append(("ℹ️ Данных пока мало", "Статистика накапливается", C_MID))

    for i, (lbl, val, color) in enumerate(insights):
        r = tr + 1 + i
        ws.merge_cells(f"A{r}:C{r}")
        lc = ws[f"A{r}"]
        lc.value, lc.font, lc.fill, lc.alignment, lc.border = (
            lbl, _font(9, True, color), _fill(C_ROW_ODD if i%2 else C_ROW_EVEN), _align("left"), _border()
        )
        ws.merge_cells(f"D{r}:H{r}")
        vc = ws[f"D{r}"]
        vc.value, vc.font, vc.fill, vc.alignment, vc.border = (
            val, _font(10, True, C_DARK), _fill(C_ROW_ODD if i%2 else C_ROW_EVEN), _align("left"), _border()
        )
        ws.row_dimensions[r].height = 22

    _widths(ws, {c: 20 for c in "ABCDEFGH"})


# ══════════════════════════════════════════════════════════════════════════════
# ЛИСТ 2 — ПОДПИСЧИКИ
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_subscribers(wb: Workbook, d: dict):
    ws = wb.create_sheet("Подписчики")

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value, c.font, c.fill, c.alignment = (
        "ДИНАМИКА ПОДПИСЧИКОВ", _font(13, True, C_WHITE), _fill(C_ACCENT), _align()
    )
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 8

    _header_row(ws, 3, ["Дата", "Мин.", "Макс.", "Изменение", "Тренд", "Накопл. рост"])

    rows = d["sub_daily"]
    if not rows:
        ws.merge_cells("A4:F4")
        c = ws["A4"]
        c.value, c.font, c.alignment = "Нет данных", _font(10, italic=True, color=C_MID), _align()
    else:
        cumul = 0
        for i, (day, mn, mx, chg) in enumerate(rows):
            r    = i + 4
            odd  = i % 2 == 0
            cumul += chg
            trend = "▲" if chg > 0 else ("▼" if chg < 0 else "→")
            _data_row(ws, r, [day, mn, mx, chg, trend, cumul], odd=odd)

            chg_c = ws.cell(r, 4)
            chg_c.font = _font(9, True, C_POSITIVE if chg > 0 else (C_NEGATIVE if chg < 0 else C_MID))

            tr_c = ws.cell(r, 5)
            tr_c.font = _font(10, True, C_POSITIVE if chg > 0 else (C_NEGATIVE if chg < 0 else C_MID))

        # Итог
        lr = len(rows) + 4
        for ci, v in enumerate(["ИТОГО", rows[0][1], rows[-1][2], d["growth"], "", d["growth"]], 1):
            c = ws.cell(lr, ci, v)
            c.font, c.fill, c.alignment = _font(9, True, C_WHITE), _fill(C_HEADER), _align()
        ws.row_dimensions[lr].height = 22

        # График
        if len(rows) > 1:
            chart = LineChart()
            chart.title  = "Подписчики"
            chart.style  = 10
            chart.height = 12
            chart.width  = 24
            data = Reference(ws, min_col=3, min_row=3, max_row=3+len(rows))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=3+len(rows)))
            chart.series[0].graphicalProperties.line.solidFill = C_ACCENT
            chart.series[0].graphicalProperties.line.width     = 20000
            ws.add_chart(chart, f"A{len(rows)+8}")

    _widths(ws, {"A": 14, "B": 12, "C": 12, "D": 16, "E": 10, "F": 16})


# ══════════════════════════════════════════════════════════════════════════════
# ЛИСТ 3 — ПОСТЫ
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_posts(wb: Workbook, d: dict):
    ws = wb.create_sheet("Посты")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value, c.font, c.fill, c.alignment = (
        "АКТИВНОСТЬ ПУБЛИКАЦИЙ", _font(13, True, C_WHITE), _fill(C_ACCENT2), _align()
    )
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 8

    _header_row(ws, 3, ["Дата", "Постов", "С медиа", "% медиа", "Просмотров", "Макс.", "Среднее"])

    rows = d["posts_daily"]
    if not rows:
        ws.merge_cells("A4:G4")
        c = ws["A4"]
        c.value, c.font, c.alignment = "Нет постов", _font(10, italic=True, color=C_MID), _align()
    else:
        for i, (day, cnt, media, tv, mv, av) in enumerate(rows):
            r   = i + 4
            mp  = f"{round(media/cnt*100)}%" if cnt else "0%"
            _data_row(ws, r, [day, cnt, media, mp, int(tv), int(mv), round(av or 0)], odd=i%2==0)

        lr = len(rows) + 4
        t_cnt = d["posts_total"]
        t_mp  = f"{round(d['media_total']/t_cnt*100)}%" if t_cnt else "0%"
        for ci, v in enumerate(["ИТОГО", t_cnt, d["media_total"], t_mp, d["views_sum"], d["views_max"], d["views_avg"]], 1):
            c = ws.cell(lr, ci, v)
            c.font, c.fill, c.alignment = _font(9, True, C_WHITE), _fill(C_HEADER), _align()
        ws.row_dimensions[lr].height = 22

        if len(rows) > 1:
            chart = BarChart()
            chart.type   = "col"
            chart.title  = "Посты по дням"
            chart.style  = 10
            chart.height = 12
            chart.width  = 24
            data = Reference(ws, min_col=2, min_row=3, max_row=3+len(rows))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=3+len(rows)))
            if chart.series:
                chart.series[0].graphicalProperties.solidFill = C_ACCENT2
            ws.add_chart(chart, f"A{len(rows)+8}")

    _widths(ws, {"A": 14, "B": 10, "C": 12, "D": 12, "E": 14, "F": 14, "G": 14})


# ══════════════════════════════════════════════════════════════════════════════
# ЛИСТ 4 — ПО ЧАСАМ
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_hourly(wb: Workbook, d: dict):
    ws = wb.create_sheet("По часам")

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value, c.font, c.fill, c.alignment = (
        "АКТИВНОСТЬ ПО ЧАСАМ UTC", _font(13, True, C_WHITE), _fill(C_NEUTRAL), _align()
    )
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 8

    _header_row(ws, 3, ["Час (UTC)", "Постов", "% от всех", "Интенсивность"])

    hmap  = {r[0]: r[1] for r in d["hourly"]}
    total = sum(hmap.values()) or 1
    mx    = max(hmap.values(), default=1)

    for h in range(24):
        cnt  = hmap.get(h, 0)
        pct  = round(cnt / total * 100, 1)
        r    = h + 4
        odd  = h % 2 == 0
        _data_row(ws, r, [f"{h:02d}:00", cnt, f"{pct}%", ""], odd=odd)

        bar_len = round(cnt / mx * 20) if mx else 0
        bc = ws.cell(r, 4, "█" * bar_len)
        intensity = min(int(cnt / mx * 200) + 55, 255) if mx else 55
        bc.font      = _font(9, color=f"5B{intensity:02X}F5")
        bc.fill      = _fill(C_ROW_ODD if odd else C_ROW_EVEN)
        bc.border    = _border()
        bc.alignment = _align("left")

    lr = 28
    for ci, v in enumerate(["ИТОГО", f"=SUM(B4:B27)", "100%", ""], 1):
        c = ws.cell(lr, ci, v)
        c.font, c.fill, c.alignment = _font(9, True, C_WHITE), _fill(C_HEADER), _align()
    ws.row_dimensions[lr].height = 22

    if d["hourly"]:
        chart = BarChart()
        chart.type   = "col"
        chart.title  = "Посты по часам"
        chart.style  = 10
        chart.height = 14
        chart.width  = 20
        data = Reference(ws, min_col=2, min_row=3, max_row=27)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=27))
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = C_NEUTRAL
        ws.add_chart(chart, "F3")

    _widths(ws, {"A": 12, "B": 10, "C": 14, "D": 24})


# ══════════════════════════════════════════════════════════════════════════════
# ГЛАВНАЯ ФУНКЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

async def generate_report(
    db_path: str,
    channel_id: int,
    channel_name: str,
    period_label: str,
    period_days: Optional[int],
) -> str:
    """Генерирует отчёт и возвращает путь к временному .xlsx файлу."""
    d  = await _fetch(db_path, channel_id, period_days)
    ts = datetime.utcnow().strftime("%d.%m.%Y %H:%M")
    wb = Workbook()

    _sheet_summary(wb, channel_name, period_label, d, ts)
    _sheet_subscribers(wb, d)
    _sheet_posts(wb, d)
    _sheet_hourly(wb, d)

    safe = "".join(c for c in channel_name if c.isalnum() or c in "-_@")[:28]
    name = f"report_{safe}_{period_label}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    path = os.path.join(tempfile.gettempdir(), name)
    wb.save(path)
    return path
